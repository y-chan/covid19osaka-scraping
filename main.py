import openpyxl
import codecs

from datetime import datetime, timedelta, timezone
from json import dumps

from typing import Dict


jst = timezone(timedelta(hours=9), 'JST')

MAIN_SUMMARY_INIT = {
    "attr": "検査実施人数",
    "value": 0,
    "children": [
        {
            "attr": "陽性患者数",
            "value": 0,
            "children": [
                {
                    "attr": "入院／入院調整中",
                    "value": 0,
                    "children": [
                        {
                            "attr": "軽症・中等症",
                            "value": 0
                        },
                        {
                            "attr": "重症",
                            "value": 0
                        }
                    ]
                },
                {
                    "attr": "退院",
                    "value": 0
                },
                {
                    "attr": "死亡",
                    "value": 0
                }
            ]
        }
    ]
}


def excel_date(num) -> datetime:
    return datetime(1899, 12, 30) + timedelta(days=num, hours=8)


def dumps_json(file_name: str, json_data: Dict) -> None:
    with codecs.open("./data/" + file_name, "w", "utf-8") as f:
        f.write(dumps(json_data, ensure_ascii=False, indent=4, separators=(',', ': ')))


class DataJson:
    def __init__(self):
        self.patients_sheet = openpyxl.load_workbook("patients.xlsx")["Sheet1"]
        self.inspections_sheet = openpyxl.load_workbook("inspections.xlsx")["モトデータ"]
        self.summary_sheet = openpyxl.load_workbook("inspections.xlsx")["総括表"]
        self.patients_count = 2
        self.inspections_count = 3
        self._data_json = {}
        self._patients_json = {}
        self._patients_summary_json = {}
        self._inspections_summary_json = {}
        self._treated_summary_json = {}
        self._main_summary_json = {}
        self.last_update = str(datetime.today().astimezone(jst).strftime("%Y/%m/%d %H:%M"))
        self.get_patients()
        self.get_inspections()

    def data_json(self):
        if not self._data_json:
            self.make_data()
        return self._data_json

    def patients_json(self) -> Dict:
        if not self._patients_json:
            self.make_patients()
        return self._patients_json

    def patients_summary_json(self) -> Dict:
        if not self._patients_summary_json:
            self.make_patients_summary()
        return self._patients_summary_json

    def inspections_summary_json(self) -> Dict:
        if not self._inspections_summary_json:
            self.make_inspections_summary()
        return self._inspections_summary_json

    def treated_summary_json(self) -> Dict:
        if not self._treated_summary_json:
            self.make_treated_summary()
        return self._treated_summary_json

    def main_summary_json(self) -> Dict:
        if not self._main_summary_json:
            self.make_main_summary()
        return self._main_summary_json

    def make_patients(self) -> None:
        self._patients_json = {
            "date": self.last_update,
            "data": []
        }
        for i in range(5, self.patients_count):
            data = {}
            release_date = excel_date(self.patients_sheet.cell(row=i, column=2).value)
            data["No"] = self.patients_sheet.cell(row=i, column=1).value
            data["リリース日"] = release_date.isoformat() + ".000Z"
            data["曜日"] = self.patients_sheet.cell(row=i, column=2).value
            data["居住地"] = self.patients_sheet.cell(row=i, column=5).value
            if not self.patients_sheet.cell(row=i, column=6).value == "―":
                data["居住地"] += " " + self.patients_sheet.cell(row=i, column=6).value
            data["年代"] = str(self.patients_sheet.cell(row=i, column=3).value) + (
                "代" if isinstance(self.patients_sheet.cell(row=i, column=3).value, int) else ""
            )
            data["性別"] = self.patients_sheet.cell(row=i, column=4).value
            data["退院"] = "〇" if "退院" in str(self.patients_sheet.cell(row=i, column=8).value) else ""
            data["date"] = release_date.strftime("%Y-%m-%d")
            self._patients_json["data"].append(data)

    def make_patients_summary(self) -> None:
        self._patients_summary_json = {
            "date": self.last_update,
            "data": []
        }

        for i in range(3, self.inspections_count):
            data = {}
            date = excel_date(self.inspections_sheet.cell(row=i, column=1).value)
            data["日付"] = date.isoformat() + ".000Z"
            data["小計"] = self.inspections_sheet.cell(row=i, column=3).value
            self._patients_summary_json["data"].append(data)

    def make_inspections_summary(self) -> None:
        self._inspections_summary_json = {
            "date": self.last_update,
            "data": []
        }
        for i in range(3, self.inspections_count):
            data = {}
            date = excel_date(self.inspections_sheet.cell(row=i, column=1).value)
            data["日付"] = date.isoformat() + ".000Z"
            data["小計"] = self.inspections_sheet.cell(row=i, column=2).value
            self._inspections_summary_json["data"].append(data)

    def make_discharges(self) -> None:
        pass

    def make_main_summary(self) -> None:
        self._main_summary_json = MAIN_SUMMARY_INIT
        all_inspections = 0
        all_patients = 0
        all_discharges = 0
        for i in range(3, self.inspections_count):
            all_inspections += self.inspections_sheet.cell(row=i, column=2).value
            all_patients += self.inspections_sheet.cell(row=i, column=3).value
            all_discharges += self.inspections_sheet.cell(row=i, column=9).value
        self._main_summary_json["value"] = all_inspections
        self._main_summary_json["children"][0]["value"] = all_patients
        self._main_summary_json["children"][0]["children"][0]["value"] = (
                all_patients - self.summary_sheet.cell(row=6, column=9).value
        )
        self._main_summary_json["children"][0]["children"][0]["children"][0]["value"] = \
            self.summary_sheet.cell(row=6, column=8).value
        self._main_summary_json["children"][0]["children"][0]["children"][1]["value"] = \
            self.summary_sheet.cell(row=6, column=7).value
        self._main_summary_json["children"][0]["children"][1]["value"] = all_discharges
        self._main_summary_json["children"][0]["children"][2]["value"] = \
            self.summary_sheet.cell(row=6, column=10).value

    def make_treated_summary(self) -> None:
        self._treated_summary_json = {
            "date": self.last_update,
            "data": []
        }

        for i in range(3, self.inspections_count):
            data = {}
            date = excel_date(self.inspections_sheet.cell(row=i, column=1).value)
            data["日付"] = date.isoformat() + ".000Z"
            data["小計"] = self.inspections_sheet.cell(row=i, column=9).value
            self._treated_summary_json["data"].append(data)

    def make_data(self) -> None:
        self._data_json = {
            "patients": self.patients_json(),
            "patients_summary": self.patients_summary_json(),
            "inspections_summary": self.inspections_summary_json(),
            "treated_summary": self.treated_summary_json(),
            "lastUpdate": self.last_update,
            "main_summary": self.main_summary_json()
        }

    def get_patients(self) -> None:
        while self.patients_sheet:
            self.patients_count += 1
            value = self.patients_sheet.cell(row=self.patients_count, column=1).value
            if not value:
                break

    def get_inspections(self) -> None:
        while self.inspections_sheet:
            self.inspections_count += 1
            value = self.inspections_sheet.cell(row=self.inspections_count, column=1).value
            if value == "計":
                break


if __name__ == '__main__':
    dumps_json("data.json", DataJson().data_json())
